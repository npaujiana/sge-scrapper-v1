"""Export service for generating Excel reports."""
import os
from datetime import datetime, date
from pathlib import Path
from typing import Optional, List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from sqlalchemy.orm import Session

from database.models import Article, SocialContent, ScrapeSession
from database.connection import get_session
from config.logging_config import get_logger
from config.settings import settings


class ExportService:
    """Service for exporting data to Excel."""

    def __init__(self):
        self.logger = get_logger()

    def export_articles_to_excel(
        self,
        output_path: Optional[str] = None,
        target_date: Optional[date] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        include_content: bool = False,
    ) -> str:
        """
        Export articles to Excel file.

        Args:
            output_path: Path for output file (default: exports/articles_YYYYMMDD_HHMMSS.xlsx)
            target_date: Filter by specific date
            start_date: Filter by date range start
            end_date: Filter by date range end
            include_content: Include full article content (can make file large)

        Returns:
            Path to the generated Excel file.
        """
        with get_session() as db:
            # Query articles
            query = db.query(Article).order_by(Article.published_at.desc())

            # Apply date filters
            if target_date:
                query = query.filter(
                    Article.published_at >= datetime.combine(target_date, datetime.min.time()),
                    Article.published_at < datetime.combine(target_date, datetime.max.time())
                )
            elif start_date and end_date:
                query = query.filter(
                    Article.published_at >= datetime.combine(start_date, datetime.min.time()),
                    Article.published_at <= datetime.combine(end_date, datetime.max.time())
                )
            elif start_date:
                query = query.filter(
                    Article.published_at >= datetime.combine(start_date, datetime.min.time())
                )
            elif end_date:
                query = query.filter(
                    Article.published_at <= datetime.combine(end_date, datetime.max.time())
                )

            articles = query.all()

            if not articles:
                self.logger.warning("No articles found for export")
                raise ValueError("No articles found with the specified filters")

            # Generate output path
            if not output_path:
                exports_dir = settings.project_root / "exports"
                exports_dir.mkdir(exist_ok=True)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_path = str(exports_dir / f"articles_{timestamp}.xlsx")

            # Create workbook
            wb = Workbook()

            # Create combined Articles + Social Contents sheet
            self._create_combined_sheet(wb, articles, include_content)

            # Create Summary sheet
            self._create_summary_sheet(wb, articles)

            # Save workbook
            wb.save(output_path)
            self.logger.info(f"Exported {len(articles)} articles to {output_path}")

            return output_path

    def _create_combined_sheet(
        self,
        wb: Workbook,
        articles: List[Article],
        include_content: bool
    ) -> None:
        """Create combined Articles + Social Contents sheet."""
        ws = wb.active
        ws.title = "Articles & Social Contents"

        # Define headers - Article info + Social Content info
        headers = [
            "No",
            "Article ID",
            "Title",
            "Article URL",
            "Category",
            "Tags",
            "Author",
            "Published Date",
            "Social Platform",
            "Social Type",
            "Social URL",
            "Social Username",
            "Social Caption",
            "Thumbnail URL",
            "Screenshot",  # New column for embedded image
        ]

        if include_content:
            headers.extend(["Subtitle", "Content (Text)"])

        # Apply header styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        # Write data - one row per social content (or one row per article if no social)
        row_num = 2
        record_num = 1

        for article in articles:
            # Format tags
            tags_str = ""
            if article.tags:
                if isinstance(article.tags, list):
                    tags_str = ", ".join(str(t) for t in article.tags)
                else:
                    tags_str = str(article.tags)

            # Base article data
            article_data = [
                article.id,
                article.title,
                article.url,
                article.category or "",
                tags_str,
                article.author_name or "",
                article.published_at.strftime("%Y-%m-%d %H:%M") if article.published_at else "",
            ]

            if include_content:
                content_data = [
                    article.subtitle or "",
                    (article.content_text or "")[:32000],
                ]
            else:
                content_data = []

            # If article has social contents, create one row per social content
            if article.social_contents:
                for sc in article.social_contents:
                    social_data = [
                        sc.platform,
                        sc.content_type,
                        sc.url or "",
                        sc.username or "",
                        (sc.caption or "")[:500],
                        sc.thumbnail_url or "",
                        "",  # Screenshot column - will add image separately
                    ]

                    row_data = [record_num] + article_data + social_data + content_data

                    for col, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_num, column=col, value=value)
                        cell.border = thin_border
                        cell.alignment = Alignment(vertical="top", wrap_text=True)

                        # Make URLs clickable
                        if col == 4 and value:  # Article URL
                            cell.hyperlink = value
                            cell.font = Font(color="0563C1", underline="single")
                        elif col == 11 and value:  # Social URL
                            cell.hyperlink = value
                            cell.font = Font(color="0563C1", underline="single")
                        elif col == 14 and value:  # Thumbnail URL
                            cell.hyperlink = value
                            cell.font = Font(color="0563C1", underline="single")

                    # Embed screenshot image if available
                    if sc.screenshot_path and Path(sc.screenshot_path).exists():
                        try:
                            img = XLImage(sc.screenshot_path)
                            # Resize to reasonable dimensions
                            img.width = 100
                            img.height = 75
                            # Place in Screenshot column (column 15 = O)
                            ws.add_image(img, f"O{row_num}")
                            # Set row height to accommodate image
                            ws.row_dimensions[row_num].height = 60
                        except Exception as e:
                            self.logger.warning(f"Failed to embed image: {e}")

                    row_num += 1
                    record_num += 1
            else:
                # Article without social content - still add one row
                social_data = ["", "", "", "", "", "", ""]  # Added empty screenshot
                row_data = [record_num] + article_data + social_data + content_data

                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

                    if col == 4 and value:  # Article URL
                        cell.hyperlink = value
                        cell.font = Font(color="0563C1", underline="single")

                row_num += 1
                record_num += 1

        # Auto-adjust column widths
        column_widths = {
            1: 6,    # No
            2: 10,   # Article ID
            3: 45,   # Title
            4: 55,   # Article URL
            5: 15,   # Category
            6: 25,   # Tags
            7: 18,   # Author
            8: 18,   # Published Date
            9: 12,   # Social Platform
            10: 12,  # Social Type
            11: 55,  # Social URL
            12: 18,  # Social Username
            13: 40,  # Social Caption
            14: 50,  # Thumbnail URL
            15: 15,  # Screenshot
        }

        if include_content:
            column_widths[16] = 40  # Subtitle
            column_widths[17] = 80  # Content

        for col, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width

        # Freeze header row
        ws.freeze_panes = "A2"

    def _create_summary_sheet(self, wb: Workbook, articles: List[Article]) -> None:
        """Create the Summary sheet."""
        ws = wb.create_sheet("Summary")

        # Calculate statistics
        total_articles = len(articles)
        total_social = sum(len(a.social_contents) for a in articles)

        # Count by category
        categories = {}
        for article in articles:
            cat = article.category or "Uncategorized"
            categories[cat] = categories.get(cat, 0) + 1

        # Count by platform
        platforms = {}
        for article in articles:
            for sc in article.social_contents:
                platforms[sc.platform] = platforms.get(sc.platform, 0) + 1

        # Styles
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Title
        ws.cell(row=1, column=1, value="Export Summary").font = title_font
        ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

        # General Statistics
        ws.cell(row=4, column=1, value="General Statistics").font = header_font
        ws.cell(row=5, column=1, value="Total Articles")
        ws.cell(row=5, column=2, value=total_articles)
        ws.cell(row=6, column=1, value="Total Social Contents")
        ws.cell(row=6, column=2, value=total_social)
        ws.cell(row=7, column=1, value="Avg Social per Article")
        ws.cell(row=7, column=2, value=round(total_social / total_articles, 2) if total_articles > 0 else 0)

        # Articles by Category
        ws.cell(row=9, column=1, value="Articles by Category").font = header_font
        row = 10
        for cat, count in sorted(categories.items(), key=lambda x: -x[1]):
            ws.cell(row=row, column=1, value=cat)
            ws.cell(row=row, column=2, value=count)
            row += 1

        # Social Contents by Platform
        ws.cell(row=row + 1, column=1, value="Social Contents by Platform").font = header_font
        row += 2
        for platform, count in sorted(platforms.items(), key=lambda x: -x[1]):
            ws.cell(row=row, column=1, value=platform)
            ws.cell(row=row, column=2, value=count)
            row += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 15

    def export_by_session(
        self,
        session_id: int,
        output_path: Optional[str] = None,
        include_content: bool = False,
    ) -> str:
        """
        Export articles from a specific scrape session.

        Args:
            session_id: The scrape session ID
            output_path: Path for output file
            include_content: Include full article content

        Returns:
            Path to the generated Excel file.
        """
        with get_session() as db:
            # Get session
            session = db.query(ScrapeSession).filter(
                ScrapeSession.id == session_id
            ).first()

            if not session:
                raise ValueError(f"Session {session_id} not found")

            if not session.target_date:
                raise ValueError(f"Session {session_id} has no target date")

            # Export articles for that date
            return self.export_articles_to_excel(
                output_path=output_path,
                target_date=session.target_date,
                include_content=include_content,
            )

    def list_exports(self) -> List[dict]:
        """List all export files in the exports directory."""
        exports_dir = settings.project_root / "exports"
        if not exports_dir.exists():
            return []

        exports = []
        for file in exports_dir.glob("*.xlsx"):
            stat = file.stat()
            exports.append({
                "filename": file.name,
                "path": str(file),
                "size_kb": round(stat.st_size / 1024, 2),
                "created_at": datetime.fromtimestamp(stat.st_ctime).strftime("%Y-%m-%d %H:%M:%S"),
            })

        return sorted(exports, key=lambda x: x["created_at"], reverse=True)
